"""生成测试用的小样本 xlsx 文件"""

from pathlib import Path

import openpyxl
import pytest

FIXTURES_DIR = Path(__file__).parent


def create_sample_xlsx(tmp_path: Path, filename: str = "test_battery.xlsx") -> Path:
    """创建一个含脉冲数据的测试 xlsx 文件"""
    filepath = tmp_path / filename
    wb = openpyxl.Workbook()

    # --- Sheet 0: Cycle ---
    ws0 = wb.active
    ws0.title = "Cycle"
    ws0.append(["Cycle#", "CycleBegin", "CycleEnd", "Charge"])  # header row
    # BTS data often has metadata in first rows
    ws0.append(["BTS_TEST_001", "", "", ""])
    ws0.append([1, "2025-06-10 08:00:00", "2025-06-10 08:30:00", 0.5])  # cycle 1
    ws0.append([2, "2025-06-10 08:30:00", "2025-06-10 09:00:00", 0.3])  # cycle 2

    # --- Sheet 1: Step ---
    ws1 = wb.create_sheet("Step")
    ws1.append(["Cycle#", "Step#", "Charge"])
    ws1.append(["BTS_TEST_001", "", ""])
    ws1.append([1, "脉冲", 0.1])
    ws1.append([1, "Charge", 0.4])
    ws1.append([2, "脉冲", 0.05])
    ws1.append([2, "Charge", 0.25])

    # --- Sheet 2: Record ---
    ws2 = wb.create_sheet("Record")
    ws2.append(["Cycle#", "Step#", "Current", "Voltage", "Charge"])
    ws2.append(["BTS_TEST_001", "", "", "", ""])

    # 模拟两段脉冲（4000mA / 2.8V）
    # 脉冲开始
    ws2.append([1, "脉冲", -4.0, 4.2, 0.0])  # 正在放电
    ws2.append([1, "脉冲", -4.0, 3.8, 0.01])  # 放电中
    ws2.append([1, "脉冲", -4.0, 2.5, 0.02])  # 放电中 → 脉冲结束点
    ws2.append([1, "Charge", 1.0, 3.0, 0.03])  # 充电，跳过
    ws2.append([2, "脉冲", -4.0, 4.1, 0.0])
    ws2.append([2, "脉冲", -4.0, 3.7, 0.01])
    ws2.append([2, "脉冲", -4.0, 2.4, 0.02])

    wb.save(filepath)
    return filepath


@pytest.fixture
def sample_xlsx(tmp_path):
    """pytest fixture: 返回小样本 xlsx 路径"""
    return create_sample_xlsx(tmp_path)
