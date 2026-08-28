from pathlib import Path

import openpyxl
import pytest


@pytest.fixture
def project_root() -> Path:
    """项目根目录路径"""
    return Path(__file__).parent.parent


@pytest.fixture
def tests_root() -> Path:
    """测试目录路径"""
    return Path(__file__).parent


@pytest.fixture
def test_data_dir() -> Path:
    """测试数据目录路径"""
    return Path(__file__).parent / "data"


def pytest_collection_modifyitems(config, items):
    """自动跳过 tests/manual/ 下的测试（除非用 --run-manual 标记运行）"""
    if config.getoption("--run-manual", default=False):
        return  # 运行所有测试

    skip_manual = pytest.mark.skip(reason="manual test, use --run-manual to run")
    for item in items:
        if "manual" in str(item.fspath):
            item.add_marker(skip_manual)


def pytest_addoption(parser):
    """添加 --run-manual 命令行选项"""
    parser.addoption(
        "--run-manual",
        action="store_true",
        default=False,
        help="run tests in tests/manual/ directory",
    )


def create_sample_xlsx(tmp_path: Path, filename: str = "test_battery.xlsx") -> Path:
    """Create a test xlsx file with pulse data"""
    filepath = tmp_path / filename
    wb = openpyxl.Workbook()

    ws0 = wb.active
    ws0.title = "Cycle"
    ws0.append(["Cycle#", "CycleBegin", "CycleEnd", "Charge"])
    ws0.append(["BTS_TEST_001", "", "", ""])
    ws0.append([1, "2025-06-10 08:00:00", "2025-06-10 08:30:00", 0.5])
    ws0.append([2, "2025-06-10 08:30:00", "2025-06-10 09:00:00", 0.3])

    ws1 = wb.create_sheet("Step")
    ws1.append(["Cycle#", "Step#", "Charge"])
    ws1.append(["BTS_TEST_001", "", ""])
    ws1.append([1, "脉冲", 0.1])
    ws1.append([1, "Charge", 0.4])
    ws1.append([2, "脉冲", 0.05])
    ws1.append([2, "Charge", 0.25])

    ws2 = wb.create_sheet("Record")
    ws2.append(["Cycle#", "Step#", "Current", "Voltage", "Charge"])
    ws2.append(["BTS_TEST_001", "", "", "", ""])
    ws2.append([1, "脉冲", -4.0, 4.2, 0.0])
    ws2.append([1, "脉冲", -4.0, 3.8, 0.01])
    ws2.append([1, "脉冲", -4.0, 2.5, 0.02])
    ws2.append([1, "Charge", 1.0, 3.0, 0.03])
    ws2.append([2, "脉冲", -4.0, 4.1, 0.0])
    ws2.append([2, "脉冲", -4.0, 3.7, 0.01])
    ws2.append([2, "脉冲", -4.0, 2.4, 0.02])

    wb.save(filepath)
    return filepath


def create_sample_xlsx_with_test_date(
    tmp_path: Path,
    filename: str = "test_date_sample.xlsx",
    test_date_value: str = "10.06.2025 - 08.07.2025",
) -> Path:
    """Create a test xlsx file with a Test Date field in the Cycle sheet

    文件名故意不含 8 位连续数字，避免触发 extract_test_date 的文件名回退，
    从而隔离验证 Test Date 单元格提取路径。
    """
    filepath = tmp_path / filename
    wb = openpyxl.Workbook()

    ws0 = wb.active
    ws0.title = "Cycle"
    ws0.append(["Test Date", test_date_value])
    ws0.append(["Cycle#", "CycleBegin", "CycleEnd", "Charge"])
    ws0.append([1, "2025-06-10 08:00:00", "2025-06-10 08:30:00", 0.5])

    ws1 = wb.create_sheet("Step")
    ws1.append(["Cycle#", "Step#", "Charge"])
    ws1.append([1, "Charge", 0.4])

    ws2 = wb.create_sheet("Record")
    ws2.append(["Cycle#", "Step#", "Current", "Voltage", "Charge"])
    ws2.append([1, "脉冲", -4.0, 4.2, 0.0])

    wb.save(filepath)
    return filepath


@pytest.fixture
def sample_xlsx_with_test_date(tmp_path):
    """pytest fixture: xlsx file containing a Test Date field"""
    return create_sample_xlsx_with_test_date(tmp_path)


@pytest.fixture
def sample_xlsx(tmp_path):
    """pytest fixture: return small sample xlsx path"""
    return create_sample_xlsx(tmp_path)
